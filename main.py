import os
import pickle
import cv2
import face_recognition
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import numpy as np

def initialize_excel_sheet(file_path, subject_name, student_ids):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = subject_name
        ws.append(['Sl no.', 'USN', 'Classes Attended'])
        for idx, usn in enumerate(student_ids, start=1):
            formula = f'=COUNTIF(D{idx + 1}:ZZ{idx + 1}, "P")'
            ws.append([idx, usn, formula])
        wb.save(file_path)
    else:
        try:
            wb = openpyxl.load_workbook(file_path)
            if subject_name not in wb.sheetnames:
                ws = wb.create_sheet(title=subject_name)
                ws.append(['Sl no.', 'USN', 'Classes Attended'])
                for idx, usn in enumerate(student_ids, start=1):
                    formula = f'=COUNTIF(D{idx + 1}:ZZ{idx + 1}, "P")'
                    ws.append([idx, usn, formula])
                wb.save(file_path)
        except Exception as e:
            print(f"Error initializing Excel file: {e}")

def initialize_main_attendance_sheet(file_path, student_ids):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance Sheet'
        ws.append(['Sl no.', 'USN', 'Total Classes Attended', 'Total Classes'])
        for idx, usn in enumerate(student_ids, start=1):
            ws.append([idx, usn, 0, 0])
        wb.save(file_path)
    else:
        try:
            wb = openpyxl.load_workbook(file_path)
            if 'Attendance Sheet' not in wb.sheetnames:
                ws = wb.create_sheet(title='Attendance Sheet')
                ws.append(['Sl no.', 'USN', 'Total Classes Attended', 'Total Classes'])
                for idx, usn in enumerate(student_ids, start=1):
                    ws.append([idx, usn, 0, 0])
                wb.save(file_path)
        except Exception as e:
            print(f"Error initializing main attendance sheet: {e}")

def get_next_sl_no(ws):
    max_sl_no = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            max_sl_no = max(max_sl_no, row[0])
    return max_sl_no + 1

def get_date_column(ws, column_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            return col
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = column_name
    return new_col

def mark_attendance(USN, class_name, subject_name, file_path, status='P'):
    try:
        today_date = datetime.now().strftime('%Y-%m-%d')
        #today_date = '2024-07-17'
        initialize_excel_sheet(file_path, subject_name, studentIds)
        wb = openpyxl.load_workbook(file_path)
        ws = wb[subject_name]

        date_col = get_date_column(ws, today_date)
        usn_col = 2

        found = False
        for row in ws.iter_rows(min_row=2, max_col=usn_col, max_row=ws.max_row):
            if row[1].value == USN:
                ws.cell(row=row[0].row, column=date_col).value = status
                found = True
                break

        if not found:
            sl_no = get_next_sl_no(ws)
            formula = '=COUNTIF(D{}:ZZ{}, "P")'.format(sl_no + 1, sl_no + 1)
            new_row = [sl_no, USN, formula] + [''] * (date_col - 4) + [status]
            ws.append(new_row)

        wb.save(file_path)

    except PermissionError:
        print(f"Permission denied: Unable to write to {file_path}. Ensure the file is not open in another program.")
    except Exception as e:
        print(f"An error occurred while marking attendance: {e}")

def update_main_attendance_sheet(file_path, subject_name, absent_students):
    try:
        today_date = datetime.now().strftime('%Y-%m-%d')
        #today_date = '2024-07-16'
        date_subject_name = f"{subject_name}_{today_date}"

        wb = openpyxl.load_workbook(file_path)
        ws = wb['Attendance Sheet']

        subject_column = get_date_column(ws, date_subject_name)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            usn = row[1].value
            cell = ws.cell(row=row[0].row, column=subject_column)
            if usn in absent_students:
                cell.value = 'A'
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                cell.value = 'P'
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            total_classes = ws.cell(row=row[0].row, column=4).value
            ws.cell(row=row[0].row, column=4).value = total_classes + 1

            total_attended = ws.cell(row=row[0].row, column=3).value
            if usn not in absent_students:
                ws.cell(row=row[0].row, column=3).value = total_attended + 1

        wb.save(file_path)
        print("Main attendance sheet updated successfully.")
    except Exception as e:
        print(f"An error occurred while updating the main attendance sheet: {e}")

def create_daily_report(file_path, class_name, subject_name, absent_students, num_classes):
    try:
        today_date = datetime.now().strftime('%Y-%m-%d')
        #today_date='2024-07-16'
        date_subject_name = f"{today_date} {subject_name}"

        wb = openpyxl.load_workbook(file_path)
        if 'Daily Report' not in wb.sheetnames:
            ws = wb.create_sheet(title='Daily Report')
            ws.append(['USN', 'Total Classes Attended', 'Total Classes'])
            for usn in studentIds:
                ws.append([usn, 0, 0])
        else:
            ws = wb['Daily Report']

            # Delete columns not corresponding to today's date
            for col in range(ws.max_column, 3, -1):
                column_header = ws.cell(row=1, column=col).value
                if column_header and today_date not in str(column_header):
                    ws.delete_cols(col)

        # Add the date and subject name as the column header for the current subject
        subject_column = get_date_column(ws, date_subject_name)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            usn = row[0].value
            cell = ws.cell(row=row[0].row, column=subject_column)
            if usn in absent_students:
                cell.value = 'A'
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                cell.value = 'P'
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            total_classes = ws.cell(row=row[0].row, column=3).value
            ws.cell(row=row[0].row, column=3).value = total_classes+1

            total_attended = ws.cell(row=row[0].row, column=2).value
            if usn not in absent_students:
                ws.cell(row=row[0].row, column=2).value = total_attended+1

        wb.save(file_path)
        print("Daily report updated successfully.")
    except Exception as e:
        print(f"An error occurred while creating the daily report: {e}")


class_name = input("Enter the class name: ")
subject_name = input("Enter the subject name: ")
file_path = f'{class_name}.xlsx'

print("Loading encoded file")
with open('EncodeFile.p', 'rb') as file:
    encodeListKnownWithIds = pickle.load(file)

encodeListKnown, studentIds = encodeListKnownWithIds
print("Encode file loaded successfully")

initialize_excel_sheet(file_path, subject_name, studentIds)
initialize_main_attendance_sheet(file_path, studentIds)

cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("Error: Could not open video capture.")
    exit()

attendance_marked = set()

while True:
    success, img = cap.read()
    if not success:
        print("Error: Failed to read frame.")
        break

    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    faceCurFrame = face_recognition.face_locations(imgS)
    encodeCurrFrame = face_recognition.face_encodings(imgS, faceCurFrame)

    for encodeFace, faceLoc in zip(encodeCurrFrame, faceCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = studentIds[matchIndex]
            if name not in attendance_marked:
                print("Known face detected:", name)
                mark_attendance(name, class_name, subject_name, file_path, 'P')
                attendance_marked.add(name)

    cv2.imshow('Webcam', img)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

# Mark absentees
absent_students = []
for student in studentIds:
    if student not in attendance_marked:
        mark_attendance(student, class_name, subject_name, file_path, 'A')
        absent_students.append(student)

# Create daily report and update main attendance sheet
num_classes = 1  # Assuming 1 class per day, adjust as needed
create_daily_report(file_path, class_name, subject_name, absent_students, num_classes)
update_main_attendance_sheet(file_path, subject_name, absent_students)